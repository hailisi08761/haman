#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import json
import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_default_template(template_path):
    """
    Generates a beautiful, professional default Excel template if the user has not uploaded one.
    This guarantees that the export function works instantly out of the box.
    """
    print(f"Creating highly polished default template at: {template_path}")
    os.makedirs(os.path.dirname(template_path), exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Template"
    
    # Enable grid lines visible
    ws.views.sheetView[0].showGridLines = True
    
    # ----------------------------------------------------
    # Styles Definition
    # ----------------------------------------------------
    font_title = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Calibri", size=13, bold=True, color="2C3E50")
    font_section_header = Font(name="Calibri", size=11, bold=True, color="000000")
    font_table_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_body_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_body = Font(name="Calibri", size=10, color="000000")
    font_meta_label = Font(name="Calibri", size=9, bold=True, color="7F8C8D")
    font_small = Font(name="Calibri", size=9, color="333333")
    
    fill_header = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid") # Dark Gray / Black
    fill_table_header = PatternFill(start_color="855C1B", end_color="855C1B", fill_type="solid") # Dark Gold / Amber
    
    thin_border_side = Side(border_style="thin", color="BDC3C7")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    thick_border_bottom_side = Side(border_style="medium", color="1A1A1A")
    thick_border_bottom = Border(bottom=thick_border_bottom_side)
    
    # Column Widths (A to I: 9 columns)
    col_widths = {
        'A': 8,   # N/M
        'B': 25,  # Pic show (Product pic / name combined)
        'C': 35,  # Size Specification
        'D': 18,  # Package size
        'E': 14,  # NW/GW
        'F': 14,  # HS Code
        'G': 10,  # Qty
        'H': 14,  # FOB Price
        'I': 16,  # Total Amount
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Row Heights
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[11].height = 26
    
    # 1. Company Header Banner (Merged A1:I1)
    ws.merge_cells("A1:I1")
    ws["A1"] = "SHANDONG HAMAN METAL PRODUCTS CO., LTD"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = fill_header
    
    ws.merge_cells("A2:I2")
    ws["A2"] = "PROFORMA / COMMERCIAL INVOICE TEMPLATE"
    ws["A2"].font = font_subtitle
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].border = thick_border_bottom

    # 2. Reference & Date Metadata (G8, G9, write to H8, H9)
    ws["G8"] = "Reference NO:"
    ws["G8"].font = font_meta_label
    ws["G8"].alignment = Alignment(horizontal="right", vertical="center")
    ws["H8"] = "PI-2026-X" # Template value
    ws["H8"].font = font_body_bold
    ws["H8"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws["G9"] = "Date:"
    ws["G9"].font = font_meta_label
    ws["G9"].alignment = Alignment(horizontal="right", vertical="center")
    ws["H9"] = datetime.datetime.now().strftime("%Y.%m.%d") # Template value (YYYY.MM.DD)
    ws["H9"].font = font_body
    ws["H9"].alignment = Alignment(horizontal="left", vertical="center")
    
    # 3. Seller Block (A4:E6) and Buyer Block (A8:E10)
    ws["A4"] = "SELLER DETAILS:"
    ws["A4"].font = font_section_header
    ws.merge_cells("B4:E4")
    ws["B4"] = "SHANDONG HAMAN METAL PRODUCTS CO., LTD"
    ws["B4"].font = font_body_bold
    
    ws["A5"] = "Address:"
    ws["A5"].font = font_meta_label
    ws.merge_cells("B5:E5")
    ws["B5"] = "Haman Industrial Zone, Jinan, Shandong, China"
    ws["B5"].font = font_small
    
    ws["A6"] = "Contact:"
    ws["A6"].font = font_meta_label
    ws.merge_cells("B6:E6")
    ws["B6"] = "Email: aimi@hamanmetal.com | Phone: +86-531-888888"
    ws["B6"].font = font_small

    # Buyer Section
    ws["A8"] = "BUYER NAME:"
    ws["A8"].font = font_section_header
    ws["A8"].alignment = Alignment(vertical="center")
    
    ws.merge_cells("B8:E8")
    ws["B8"] = "Default Buyer Company Ltd." # Template value (to be rewritten)
    ws["B8"].font = font_body_bold
    ws["B8"].alignment = Alignment(vertical="center")
    ws["B8"].border = thin_border
    
    ws["A9"] = "DELIVERY ADDR:"
    ws["A9"].font = font_meta_label
    ws["A9"].alignment = Alignment(vertical="center")
    
    ws.merge_cells("B9:E10")
    ws["B9"] = "123 Shipping Lane, Port City, USA" # Template value (to be rewritten)
    ws["B9"].font = font_small
    ws["B9"].alignment = Alignment(vertical="top", wrap_text=True)
    ws["B9"].border = thin_border

    # 4. Products Table Header (Row 11 - 9 columns)
    headers = [
        "N/M", 
        "Pic show",   # Column B (Combined Item Photo & Pic show)
        "Size Specification (mm)", # Column C
        "Package size (cm)",       # Column D
        "N.W / G.W (kg)",          # Column E
        "HS Code",                 # Column F
        "Qty",                     # Column G
        "FOB Price",               # Column H
        "Total Amount"             # Column I
    ]
    for col_idx, h_text in enumerate(headers, start=1):
        cell_ref = f"{get_column_letter(col_idx)}11"
        ws[cell_ref] = h_text
        ws[cell_ref].font = font_table_header
        ws[cell_ref].fill = fill_table_header
        ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[cell_ref].border = thin_border

    # 5. Placeholders for Row 12 (Dynamic Product Starting Row)
    for col_idx in range(1, 10):
        cell_ref = f"{get_column_letter(col_idx)}12"
        ws[cell_ref].border = thin_border

    # 6. Save Template
    wb.save(template_path)
    print("Default template successfully written.")


def fill_excel_template(data_path, template_path, output_path):
    """
    Reads JSON data from data_path, loads the Excel template at template_path,
    populates the values dynamically at precise coordinates, and saves as output_path.
    """
    # Force clean template path conforming to user guidelines
    clean_template_path = "./templates/Proforma invoice-Aimi-Firepit.xlsx"
    
    if not os.path.exists(clean_template_path):
        create_default_template(clean_template_path)
        
    try:
        # Load fresh copy every single time to prevent in-memory contamination
        wb = load_workbook(clean_template_path)
        ws = wb.active
    except Exception as e:
        print(f"Error loading Excel template: {e}", file=sys.stderr)
        create_default_template(clean_template_path)
        wb = load_workbook(clean_template_path)
        ws = wb.active

    # Enable grid lines explicitly to ensure spreadsheet is readable
    ws.views.sheetView[0].showGridLines = True

    # Clear rows 12 to 35 (the products table area) of any historical residues or layout defects
    for r in range(12, 36):
        ws.row_dimensions[r].height = None
        for col_idx in range(1, 10):
            cell = ws.cell(row=r, column=col_idx)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()

    # Read data JSON
    with open(data_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Extract data fields
    buyer_name = data.get("buyerName", "")
    buyer_address = data.get("buyerAddress", "")
    reference_no = data.get("referenceNo", "")
    
    # DYNAMIC DATE COMPLIANCE (YYYY.MM.DD)
    date_val = data.get("date")
    if not date_val:
        date_val = datetime.datetime.now().strftime("%Y.%m.%d")
    else:
        # Standardize dashes/slashes to dots
        date_val = str(date_val).replace("-", ".").replace("/", ".")
        
    products = data.get("products", [])
    document_type = data.get("documentType", "PI")
    currency = data.get("currency", "USD")
    
    # Shipping metadata
    shipping_marks = data.get("shippingMarks", "N/M")
    port_of_loading = data.get("portOfLoading", "Qingdao, China")
    port_of_destination = data.get("portOfDestination", "")
    packing = data.get("packing", "Standard seaworthy packaging")
    total_cbm = data.get("totalCbm", "")
    total_weight = data.get("totalWeight", "")
    
    # Financials
    deposit_percent = data.get("depositPercent", 30)
    balance_percent = data.get("balancePercent", 70)
    deposit_amount = data.get("depositAmount", 0)
    balance_amount = data.get("balanceAmount", 0)
    grand_total = data.get("grandTotal", 0)
    
    # Bank Details
    bank_info = data.get("bankInfo", {})

    # Define reusable fonts & fills
    font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_regular = Font(name="Calibri", size=10, color="000000")
    font_mono = Font(name="Consolas", size=9, color="000000")
    font_meta_label = Font(name="Calibri", size=9, bold=True, color="7F8C8D")
    font_small = Font(name="Calibri", size=9, color="333333")
    
    fill_zebra = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid")
    fill_totals = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="BDC3C7")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom_side = Side(border_style="double", color="1A1A1A")
    thick_top_side = Side(border_style="medium", color="1A1A1A")
    totals_border = Border(top=thick_top_side, bottom=double_bottom_side, left=thin_border_side, right=thin_border_side)

    # Multi-currency support symbol
    currency_symbols = {
        "USD": "$",
        "GBP": "£",
        "EUR": "€"
    }
    symbol = currency_symbols.get(currency, "$")
    num_format = f'{symbol}#,##0.00'

    # Update Table Header with Currency context
    ws["H11"] = f"FOB Price ({currency})"
    ws["I11"] = f"Total ({currency})"

    # WRITE METADATA & BUYER INFO (PRECISE SINGLE CELL COORDINATES)
    ws["B8"] = buyer_name
    ws["B8"].font = font_bold
    ws["B8"].alignment = Alignment(vertical="center", wrap_text=True)

    ws["B9"] = buyer_address
    ws["B9"].font = font_regular
    ws["B9"].alignment = Alignment(vertical="top", wrap_text=True)

    # Reference Number at H8 or I8 (Aligned at top-right of table)
    ws["G8"] = f"{document_type} REFERENCE NO:"
    ws["G8"].font = font_meta_label
    ws["G8"].alignment = Alignment(horizontal="right", vertical="center")
    ws["H8"] = reference_no
    ws["H8"].font = font_bold
    ws["H8"].alignment = Alignment(horizontal="left", vertical="center")

    ws["G9"] = "Date:"
    ws["G9"].font = font_meta_label
    ws["G9"].alignment = Alignment(horizontal="right", vertical="center")
    ws["H9"] = date_val
    ws["H9"].font = font_regular
    ws["H9"].alignment = Alignment(horizontal="left", vertical="center")

    # Dynamically change Title depending on Document Type (PI vs CI)
    ws["A2"] = f"{'PROFORMA' if document_type == 'PI' else 'COMMERCIAL'} INVOICE"
    
    # -------------------------------------------------------------------------
    # WRITE PRODUCT ROWS WITH REARRANGED COLUMNS (START FROM ROW 12)
    # -------------------------------------------------------------------------
    start_row = 12
    current_row = start_row

    for idx, p in enumerate(products):
        nw_val = p.get("nw", "")
        gw_val = p.get("gw", "")
        nw_gw_stacked = f"N.W: {nw_val}\nG.W: {gw_val}" if (nw_val or gw_val) else ""

        # Coordinates for product cells in current row (9 Columns: A to I)
        cell_nm = f"A{current_row}"      # A: N/M
        cell_pic = f"B{current_row}"     # B: Pic show (Combined Pic show & Item Photo)
        cell_size = f"C{current_row}"    # C: Size Specification (mm)
        cell_pkg = f"D{current_row}"     # D: Package size (cm)
        cell_weight = f"E{current_row}"  # E: N.W / G.W (kg)
        cell_hs = f"F{current_row}"      # F: HS Code
        cell_qty = f"G{current_row}"     # G: Qty
        cell_fob = f"H{current_row}"     # H: FOB Price
        cell_total = f"I{current_row}"    # I: Total Amount

        # Fill values
        ws[cell_nm] = idx + 1
        
        # Merge Item Photo & Pic show: write "[Item Photo]\n{pic_name}" to Column B
        pic_name = p.get("pic_name", "")
        ws[cell_pic] = f"[Item Photo]\n{pic_name}"
        
        ws[cell_size] = p.get("size_mm", "")
        ws[cell_pkg] = p.get("package_size", "")
        ws[cell_weight] = nw_gw_stacked
        ws[cell_hs] = p.get("hs_code", "")
        
        # Qty & Unit integration
        qty = int(p.get("qty", 0))
        unit = p.get("unit", "set")
        ws[cell_qty] = f"{qty} {unit}"
        
        fob_price = float(p.get("fob_usd", 0.0))
        row_total = qty * fob_price
        
        ws[cell_fob] = fob_price
        ws[cell_total] = row_total

        # Set Number Formatting
        ws[cell_fob].number_format = num_format
        ws[cell_total].number_format = num_format

        # Styling cells in the current row
        is_even = (idx % 2 == 0)
        row_fill = fill_zebra if is_even else PatternFill(fill_type=None)

        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            cell = ws[f"{col_letter}{current_row}"]
            cell.border = thin_border
            if row_fill.fill_type:
                cell.fill = row_fill
            
            # Alignments & Fonts
            if col_letter == 'A':
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = font_bold
            elif col_letter == 'B':
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.font = font_bold
            elif col_letter == 'G':
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = font_bold
            elif col_letter in ['E', 'D', 'F']:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.font = font_mono
            elif col_letter in ['H', 'I']:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = font_bold if col_letter == 'I' else font_regular
            else: # Column C: Size Specification
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.font = font_regular

        # Adjust row height based on size_mm specs length to prevent cell clipping
        lines = len(p.get("size_mm", "").split('\n'))
        ws.row_dimensions[current_row].height = max(55, lines * 14 + 15)
        
        current_row += 1

    # -------------------------------------------------------------------------
    # WRITE GRAND TOTALS ROW (DYNAMICALLY APPENDED IMMEDIATELY AFTER PRODUCTS)
    # -------------------------------------------------------------------------
    totals_row = current_row
    ws.row_dimensions[totals_row].height = 24
    
    # Merge column A to F for the "GRAND TOTAL" label
    ws.merge_cells(f"A{totals_row}:F{totals_row}")
    total_label_cell = ws[f"A{totals_row}"]
    total_label_cell.value = f"GRAND TOTAL (FOB {currency}):"
    total_label_cell.font = Font(name="Calibri", size=10, bold=True, color="855C1B")
    total_label_cell.alignment = Alignment(horizontal="right", vertical="center")
    
    # Calculate Sums
    total_qty = sum(int(p.get("qty", 0)) for p in products)
    
    # Qty Sum in Column G
    qty_cell = ws[f"G{totals_row}"]
    qty_cell.value = total_qty
    qty_cell.font = font_bold
    qty_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # FOB Header spacer column H
    fob_cell = ws[f"H{totals_row}"]
    fob_cell.value = ""
    
    # Total Sum in Column I
    total_sum_cell = ws[f"I{totals_row}"]
    total_sum_cell.value = grand_total
    total_sum_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    total_sum_cell.fill = PatternFill(start_color="855C1B", end_color="855C1B", fill_type="solid") # Golden Amber total box
    total_sum_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_sum_cell.number_format = num_format

    # Apply borders to the Totals row (A to I)
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        c = ws[f"{col_letter}{totals_row}"]
        c.border = totals_border
        if col_letter != 'I':
            c.fill = fill_totals

    current_row += 2 # Add spacer row before terms and bank information

    # -------------------------------------------------------------------------
    # CONDITIONAL CI HEADERS & CARGO DECLARATION OR PI BANK / PAYMENT TERMS
    # -------------------------------------------------------------------------
    if document_type == 'CI':
        # --- Commercial Invoice Specific Info ---
        ws.row_dimensions[current_row].height = 20
        ws.merge_cells(f"A{current_row}:I{current_row}")
        section_cell = ws[f"A{current_row}"]
        section_cell.value = "SHIPPING & CLEARANCE SUMMARY"
        section_cell.font = font_bold
        section_cell.fill = fill_totals
        section_cell.border = Border(bottom=Side(style="thin", color="000000"))
        
        current_row += 1
        ws.row_dimensions[current_row].height = 20
        ws[f"A{current_row}"] = "Shipping Marks:"
        ws[f"A{current_row}"].font = font_meta_label
        ws.merge_cells(f"B{current_row}:D{current_row}")
        ws[f"B{current_row}"] = shipping_marks
        ws[f"B{current_row}"].font = font_bold
        
        ws[f"E{current_row}"] = "Port of Loading:"
        ws[f"E{current_row}"].font = font_meta_label
        ws.merge_cells(f"F{current_row}:I{current_row}")
        ws[f"F{current_row}"] = port_of_loading
        ws[f"F{current_row}"].font = font_bold
        
        current_row += 1
        ws.row_dimensions[current_row].height = 20
        ws[f"A{current_row}"] = "Destination Port:"
        ws[f"A{current_row}"].font = font_meta_label
        ws.merge_cells(f"B{current_row}:D{current_row}")
        ws[f"B{current_row}"] = port_of_destination
        ws[f"B{current_row}"].font = font_bold
        
        ws[f"E{current_row}"] = "Total Volume CBM:"
        ws[f"E{current_row}"].font = font_meta_label
        ws.merge_cells(f"F{current_row}:I{current_row}")
        ws[f"F{current_row}"] = total_cbm
        ws[f"F{current_row}"].font = font_bold
        
        current_row += 1
        ws.row_dimensions[current_row].height = 20
        ws[f"A{current_row}"] = "Packing Terms:"
        ws[f"A{current_row}"].font = font_meta_label
        ws.merge_cells(f"B{current_row}:D{current_row}")
        ws[f"B{current_row}"] = packing
        ws[f"B{current_row}"].font = font_regular
        
        ws[f"E{current_row}"] = "Total Wt (NW/GW):"
        ws[f"E{current_row}"].font = font_meta_label
        ws.merge_cells(f"F{current_row}:I{current_row}")
        ws[f"F{current_row}"] = total_weight
        ws[f"F{current_row}"].font = font_bold
        
    else:
        # --- Proforma Invoice Specific Info (Deposit and Bank Details) ---
        # 1. Payment Terms Section
        ws.row_dimensions[current_row].height = 20
        ws.merge_cells(f"A{current_row}:I{current_row}")
        section_cell = ws[f"A{current_row}"]
        section_cell.value = "PAYMENT & DEPOSIT SPLIT TERMS"
        section_cell.font = font_bold
        section_cell.fill = fill_totals
        section_cell.border = Border(bottom=Side(style="thin", color="000000"))
        
        current_row += 1
        ws.row_dimensions[current_row].height = 18
        ws[f"A{current_row}"] = f"{deposit_percent}% Deposit Amount:"
        ws[f"A{current_row}"].font = font_meta_label
        ws.merge_cells(f"B{current_row}:D{current_row}")
        ws[f"B{current_row}"] = deposit_amount
        ws[f"B{current_row}"].font = font_bold
        ws[f"B{current_row}"].number_format = num_format
        
        ws[f"E{current_row}"] = f"{balance_percent}% Balance Amount:"
        ws[f"E{current_row}"].font = font_meta_label
        ws.merge_cells(f"F{current_row}:I{current_row}")
        ws[f"F{current_row}"] = balance_amount
        ws[f"F{current_row}"].font = font_bold
        ws[f"F{current_row}"].number_format = num_format
        
        current_row += 1
        ws.row_dimensions[current_row].height = 18
        ws[f"A{current_row}"] = "Instruction:"
        ws[f"A{current_row}"].font = font_meta_label
        ws.merge_cells(f"B{current_row}:I{current_row}")
        ws[f"B{current_row}"] = f"Prepay {deposit_percent}% deposit via {currency}, and pay the remaining {balance_percent}% balance before shipment."
        ws[f"B{current_row}"].font = font_small
        
        # 2. Banking Details Section
        current_row += 2
        ws.row_dimensions[current_row].height = 20
        ws.merge_cells(f"A{current_row}:I{current_row}")
        bank_section_cell = ws[f"A{current_row}"]
        bank_section_cell.value = f"SETTLEMENT BANK DETAILS (银行收款路径 - {currency} ACCOUNT)"
        bank_section_cell.font = font_bold
        bank_section_cell.fill = fill_totals
        bank_section_cell.border = Border(bottom=Side(style="thin", color="000000"))
        
        bank_fields = [
            ("BENEFICIARY BANK", bank_info.get("beneficiaryBank", "")),
            ("BENEFICIARY BANK ADDRESS", bank_info.get("beneficiaryBankAddress", "")),
            ("SWIFT CODE (SWIFT)", bank_info.get("swiftCode", "")),
            ("ACCOUNT NUMBER", bank_info.get("accountNumber", "")),
            ("BENEFICIARY NAME", bank_info.get("sellerCompany", "")),
            ("BENEFICIARY ADDRESS", bank_info.get("sellerAddress", "")),
        ]
        
        for label, val in bank_fields:
            current_row += 1
            ws.row_dimensions[current_row].height = 18
            ws[f"A{current_row}"] = label + ":"
            ws[f"A{current_row}"].font = font_meta_label
            ws.merge_cells(f"B{current_row}:I{current_row}")
            ws[f"B{current_row}"] = val
            ws[f"B{current_row}"].font = font_bold if label in ["SWIFT CODE (SWIFT)", "ACCOUNT NUMBER", "BENEFICIARY NAME"] else font_regular
            ws[f"B{current_row}"].alignment = Alignment(horizontal="left", vertical="center")

    # -------------------------------------------------------------------------
    # REMARKS SECTION
    # -------------------------------------------------------------------------
    remarks = data.get("remarks", [])
    if remarks:
        current_row += 2
        ws.row_dimensions[current_row].height = 20
        ws.merge_cells(f"A{current_row}:I{current_row}")
        rem_section = ws[f"A{current_row}"]
        rem_section.value = "REMARKS / ADDITIONAL TERMS"
        rem_section.font = font_bold
        rem_section.fill = fill_totals
        rem_section.border = Border(bottom=Side(style="thin", color="000000"))
        
        for idx, rem in enumerate(remarks):
            current_row += 1
            ws.row_dimensions[current_row].height = 18
            ws[f"A{current_row}"] = f"{(idx+1)}."
            ws[f"A{current_row}"].font = font_bold
            ws[f"A{current_row}"].alignment = Alignment(horizontal="center")
            ws.merge_cells(f"B{current_row}:I{current_row}")
            ws[f"B{current_row}"] = rem
            ws[f"B{current_row}"].font = font_small
            ws[f"B{current_row}"].alignment = Alignment(horizontal="left", vertical="center")

    # -------------------------------------------------------------------------
    # SIGNATURE BLOCK
    # -------------------------------------------------------------------------
    current_row += 3
    ws.row_dimensions[current_row].height = 18
    ws.merge_cells(f"A{current_row}:D{current_row}")
    ws[f"A{current_row}"] = "Approved by (Buyer):"
    ws[f"A{current_row}"].font = font_meta_label
    
    ws.merge_cells(f"G{current_row}:I{current_row}")
    ws[f"G{current_row}"] = "Authorized Signature (Seller):"
    ws[f"G{current_row}"].font = font_meta_label
    
    current_row += 2
    ws.row_dimensions[current_row].height = 30
    ws.merge_cells(f"A{current_row}:D{current_row}")
    ws[f"A{current_row}"] = "__________________________"
    ws[f"A{current_row}"].font = font_regular
    ws[f"A{current_row}"].alignment = Alignment(vertical="bottom")
    
    ws.merge_cells(f"G{current_row}:I{current_row}")
    ws[f"G{current_row}"] = "__________________________"
    ws[f"G{current_row}"].font = font_regular
    ws[f"G{current_row}"].alignment = Alignment(vertical="bottom")

    # Save output Excel workbook
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Excel report saved successfully to: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python3 export_excel.py <data_path.json> <template_path.xlsx> <output_path.xlsx>")
        sys.exit(1)
        
    dp = sys.argv[1]
    tp = sys.argv[2]
    op = sys.argv[3]
    
    fill_excel_template(dp, tp, op)
